#bfdev

import pyautogui 
import pyperclip 
import openpyxl 
import time 

# --- CONFIGURAÇÃO --- 
pyautogui.PAUSE = 1 
pyautogui.FAILSAFE = True 

# Nome do arquivo Excel
ARQUIVO_EXCEL = "input_dados_fornecedores.xlsx" 
PLANILHA_EXCEL = "Plan1" 

# --- INÍCIO DA AUTOMAÇÃO --- 
# PREPARE O AMBIENTE
pyautogui.alert("A automação vai começar. Deixe a janela do SAP visível. Pressione OK para iniciar.") 

try: 
    workbook = openpyxl.load_workbook(ARQUIVO_EXCEL) 
    sheet = workbook[PLANILHA_EXCEL] 
except FileNotFoundError: 
    pyautogui.alert(f"Erro: Arquivo '{ARQUIVO_EXCEL}' não encontrado. Verifique se ele está na mesma pasta do script.") 
    exit() 

# loop principal começa aqui
for linha in range(2, sheet.max_row + 1): 
    codigo_erp = sheet.cell(row=linha, column=2).value 
    
    if codigo_erp is None: 
        break # Fim da lista
        
    pyperclip.copy(str(codigo_erp)) 

    try: 
        # 1. Localizar e Clicar no botões/campos Fornecedor do SAP (Via Imagem)
        # OBS: As imagens (.png) precisam estar na pasta do projeto
        campo_fornecedor_pos = pyautogui.locateCenterOnScreen('img_campo_fornecedor.png', confidence=0.9) 
        if campo_fornecedor_pos is None: 
            raise Exception("Não foi possível encontrar o campo 'Fornecedor' na tela do SAP.") 
        
        pyautogui.click(campo_fornecedor_pos.x + 150, campo_fornecedor_pos.y) 
        pyautogui.hotkey('ctrl', 'a') 
        pyautogui.hotkey('ctrl', 'v') 

        # 2. Acessar a tela de pagamentos 
        pyautogui.press('enter') 
        time.sleep(2) 

        # 3. Extração 1:  Titular da Conta
        label_titular_pos = pyautogui.locateCenterOnScreen('img_label_titular.png', confidence=0.9) 
        if label_titular_pos is None: 
            raise Exception("Campo 'Titular' não encontrado.") 
            
        pyautogui.click(label_titular_pos.x, label_titular_pos.y + 25) 
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.hotkey('ctrl', 'c') 
        titular_cta = pyperclip.paste() 

        # 4. Extração 2: Chave do Banco
        label_chave_pos = pyautogui.locateCenterOnScreen('img_label_chave_banco.png', confidence=0.9) 
        if label_chave_pos is None: 
            raise Exception("Campo 'Chave do Banco' não encontrado.") 
            
        pyautogui.click(label_chave_pos.x, label_chave_pos.y + 25)
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.hotkey('ctrl', 'c') 
        chave_banco = pyperclip.paste() 

        # 5. Extração 3: Conta Bancária
        label_conta_pos = pyautogui.locateCenterOnScreen('img_label_conta.png', confidence=0.9)
        if label_conta_pos is None:
            raise Exception("Campo 'Conta Bancária' não encontrado.")
        
        pyautogui.click(label_conta_pos.x, label_conta_pos.y + 25)
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.hotkey('ctrl', 'c')
        conta_bancaria = pyperclip.paste()

        # 6. Extração 4: CC
        label_cc_pos = pyautogui.locateCenterOnScreen('img_label_cc.png', confidence=0.9)
        if label_cc_pos is None:
            raise Exception("Campo 'CC' não encontrado.")

        pyautogui.click(label_cc_pos.x, label_cc_pos.y + 25)
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.hotkey('ctrl', 'c')
        codigo_cc = pyperclip.paste()

        # 7. Extração 5: Recebedor Divergente
        label_recebedor_pos = pyautogui.locateCenterOnScreen('img_label_recebedor.png', confidence=0.85)
        if label_recebedor_pos is None:
            raise Exception("Campo 'Recebedor' não encontrado.")
            
        pyautogui.click(label_recebedor_pos.x + 150, label_recebedor_pos.y)
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.hotkey('ctrl', 'c')
        recebedor_dif = pyperclip.paste()
        
        # --- GRAVAÇÃO DOS DADOS NO EXCEL ---
        sheet.cell(row=linha, column=15).value = titular_cta 
        sheet.cell(row=linha, column=16).value = chave_banco 
        sheet.cell(row=linha, column=17).value = recebedor_dif
        sheet.cell(row=linha, column=18).value = conta_bancaria
        sheet.cell(row=linha, column=19).value = codigo_cc

        # Retornar para a tela inicial usando o atalho F3
        pyautogui.press('f3') 
        time.sleep(2) 
        
    except Exception as e: 
        pyautogui.alert(f"Ocorreu um erro na linha {linha}:\n{e}\nParando execução.") 
        break 

# Salva arquivo final com todos os dados extraidos
try:
    workbook.save(ARQUIVO_EXCEL)
    pyautogui.alert("Automação concluída com sucesso!")
except Exception as e:
    pyautogui.alert(f"Erro ao salvar arquivo:\n{e}")

#bfdev